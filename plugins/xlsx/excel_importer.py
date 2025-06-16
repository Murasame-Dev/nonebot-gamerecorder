#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import glob
from openpyxl import load_workbook
from typing import List, Optional
from .config import Config
from .database import DatabaseManager

class ExcelImporter:
    """Excel文件导入工具"""
    
    def __init__(self):
        self.config = Config()
        self.db_manager = DatabaseManager()
    
    def get_excel_files(self) -> List[str]:
        """获取目标文件夹中的所有xlsx文件"""
        pattern = os.path.join(self.config.excel_folder, "*.xlsx")
        excel_files = glob.glob(pattern)
        # 过滤掉临时文件（以~$开头的文件）
        excel_files = [f for f in excel_files if not os.path.basename(f).startswith('~$')]
        return excel_files
    
    def get_excel_file_by_name(self, filename: str) -> Optional[str]:
        """根据文件名查找Excel文件"""
        excel_files = self.get_excel_files()
        
        # 精确匹配
        for filepath in excel_files:
            if os.path.basename(filepath) == filename:
                return filepath
            # 不带扩展名的匹配
            if os.path.basename(filepath) == f"{filename}.xlsx":
                return filepath
            # 文件名部分匹配
            if filename in os.path.basename(filepath):
                return filepath
        
        return None
    
    def read_excel_data(self, file_path: str) -> List[List[str]]:
        """读取Excel文件数据"""
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            if ws is None:
                raise ValueError("Excel文件格式错误")            
            data = []
            for row in range(1, ws.max_row + 1):
                row_data = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data.append(str(cell_value))
                    else:
                        row_data.append("")
                # 只添加非空行（至少A列有数据）
                if row_data and row_data[0].strip():
                    data.append(row_data)
            
            return data
            
        except Exception as e:
            raise ValueError(f"读取Excel文件失败: {str(e)}")
    
    def import_excel_file(self, filename: str) -> str:
        """导入Excel文件到数据库"""
        # 查找文件
        file_path = self.get_excel_file_by_name(filename)
        if not file_path:
            available_files = [os.path.basename(f) for f in self.get_excel_files()]
            return f"❌ 未找到文件: {filename}\n可用文件: {', '.join(available_files)}"
        
        # 获取游戏名（从文件名提取）
        game_name = os.path.basename(file_path)
        if game_name.endswith('.xlsx'):
            game_name = game_name[:-5]
        
        try:
            # 读取Excel数据
            excel_data = self.read_excel_data(file_path)
            
            if not excel_data:
                return f"❌ 文件 {filename} 没有有效数据"
            
            # 使用对比导入功能
            result = self.db_manager.import_from_excel_data_with_comparison(game_name, excel_data)
            
            # 构建返回消息
            message = f"✅ 成功导入文件: {filename}\n"
            message += f"🎮 游戏: {result['game_name']}\n"
            
            if result['is_existing_game']:
                message += f"📊 数据库对比结果:\n"
                message += f"  • 导入前记录数: {result['records_before']}\n"
                message += f"  • 导入后记录数: {result['records_after']}\n"
                message += f"  • 新增记录数: {result['new_records']}\n"
                message += f"  • 处理记录数: {result['imported_count']}"
                
                if result['new_records'] == 0:
                    message += f"\n💡 提示: 没有新增记录，可能数据已存在"
                elif result['new_records'] != result['imported_count']:
                    message += f"\n💡 提示: 部分记录可能已存在或重复"
            else:
                message += f"📝 新建游戏，导入记录数: {result['imported_count']}"
            
            return message
            
        except Exception as e:
            return f"❌ 导入失败: {str(e)}"
    
    def import_excel(self, file_path: str) -> str:
        """通用Excel导入方法，支持任意路径的Excel文件"""
        if not os.path.exists(file_path):
            return f"❌ 文件不存在: {file_path}"
        
        # 获取游戏名（从文件名提取）
        game_name = os.path.basename(file_path)
        if game_name.endswith('.xlsx'):
            game_name = game_name[:-5]
        elif game_name.endswith('.xls'):
            game_name = game_name[:-4]
        
        try:
            # 读取Excel数据
            excel_data = self.read_excel_data(file_path)
            
            if not excel_data:
                return f"❌ 文件 {os.path.basename(file_path)} 没有有效数据"
            
            # 使用对比导入功能
            result = self.db_manager.import_from_excel_data_with_comparison(game_name, excel_data)
            
            # 构建返回消息
            message = f"✅ 成功导入文件: {os.path.basename(file_path)}\n"
            message += f"🎮 游戏: {result['game_name']}\n"
            
            if result['is_existing_game']:
                message += f"📊 数据库对比结果:\n"
                message += f"  • 导入前记录数: {result['records_before']}\n"
                message += f"  • 导入后记录数: {result['records_after']}\n"
                message += f"  • 新增记录数: {result['new_records']}\n"
                message += f"  • 处理记录数: {result['imported_count']}"
                
                if result['new_records'] == 0:
                    message += f"\n💡 提示: 没有新增记录，可能数据已存在"
                elif result['new_records'] != result['imported_count']:
                    message += f"\n💡 提示: 部分记录可能已存在或重复"
            else:
                message += f"📝 新建游戏，导入记录数: {result['imported_count']}"
            
            return message
            
        except Exception as e:
            return f"❌ 导入失败: {str(e)}"
    
    def list_available_files(self) -> str:
        """列出可用的Excel文件"""
        excel_files = self.get_excel_files()
        
        if not excel_files:
            return f"❌ 在目录 {self.config.excel_folder} 中未找到Excel文件"
        
        file_list = []
        for filepath in excel_files:
            filename = os.path.basename(filepath)
            game_name = filename[:-5] if filename.endswith('.xlsx') else filename
            file_list.append(f"• {filename} ({game_name})")
        
        return f"📁 可用的Excel文件:\n" + "\n".join(file_list)
