#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from typing import List, Tuple, Optional, Dict
from datetime import datetime
from .config import Config
from .database import DatabaseManager

class ExcelExporter:
    """Excel文件导出工具"""
    
    def __init__(self):
        self.config = Config()
        self.db_manager = DatabaseManager()
          # 定义样式
        self.blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色填充
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.bold_font = Font(bold=True)
    
    def get_game_data(self, game_name: str) -> Optional[Dict]:
        """获取游戏的完整数据"""
        conn = sqlite3.connect(self.db_manager.db_path)
        cursor = conn.cursor()
        
        # 获取游戏ID
        cursor.execute('SELECT id FROM games WHERE name = ?', (game_name,))
        game_result = cursor.fetchone()
        if not game_result:
            conn.close()
            return None
        
        game_id = game_result[0]
        
        # 获取所有用户及其记录
        cursor.execute('''
            SELECT u.id, u.name, u.cycle, u.is_completed
            FROM users u
            WHERE u.game_id = ?
            ORDER BY u.name, u.cycle
        ''', (game_id,))
        
        users = cursor.fetchall()
        
        # 获取每个用户的记录
        user_data = []
        for user_id, user_name, cycle, is_completed in users:
            cursor.execute('''
                SELECT record_date, count
                FROM records
                WHERE user_id = ?
                ORDER BY id
            ''', (user_id,))
            
            records = cursor.fetchall()
            user_data.append({
                'id': user_id,
                'name': user_name,
                'cycle': cycle,
                'is_completed': bool(is_completed),
                'records': records
            })
        
        conn.close()
        
        return {
            'game_name': game_name,
            'game_id': game_id,
            'users': user_data        }
    
    def create_excel_file(self, game_data: Dict) -> str:
        """根据游戏数据创建Excel文件"""
        wb = Workbook()
        ws = wb.active
        ws.title = "代肝记录"
        
        game_name = game_data['game_name']
        users = game_data['users']
        
        if not users:
            # 如果没有用户数据，创建空表格
            # 不设置表头，直接设置行高
            ws.row_dimensions[1].height = self.config.row_height
        else:
            # 组织数据：将同名用户的不同周期合并显示
            organized_users = {}
            for user in users:
                base_name = user['name']
                cycle = user['cycle']
                
                if base_name not in organized_users:
                    organized_users[base_name] = []
                
                organized_users[base_name].append(user)
            
            # 写入数据
            current_row = 1
            
            for base_name, user_cycles in organized_users.items():
                # 按周期排序
                user_cycles.sort(key=lambda x: x['cycle'])
                
                for user in user_cycles:
                    cycle = user['cycle']
                    is_completed = user['is_completed']
                    records = user['records']
                    
                    # 确定用户名显示格式
                    if cycle == 1:
                        display_name = base_name
                    else:
                        display_name = f"{base_name}({cycle})"                    # 写入用户名
                    ws.cell(row=current_row, column=1, value=display_name)
                    ws.cell(row=current_row, column=1).alignment = self.center_alignment
                    ws.cell(row=current_row, column=1).fill = self.yellow_fill  # A列用户名设置为黄色
                    ws.row_dimensions[current_row].height = self.config.row_height
                    
                    # 写入记录
                    col = 2
                    for record_date, count in records:
                        record_value = f"{record_date}_{count}"
                        ws.cell(row=current_row, column=col, value=record_value)
                        ws.cell(row=current_row, column=col).alignment = self.center_alignment
                        col += 1
                    
                    # 如果周期已完成，设置记录列（除A列外）为蓝色背景
                    if is_completed:
                        for c in range(2, col):  # 从第2列开始，跳过A列用户名
                            ws.cell(row=current_row, column=c).fill = self.blue_fill
                    
                    current_row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = self.config.name_column_width
        
        # 保存文件
        export_folder = os.path.join(self.config.excel_folder, "exports")
        os.makedirs(export_folder, exist_ok=True)
        
        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime("%m-%d-%H%M")
        filename = f"{game_name}_export_{timestamp}.xlsx"
        file_path = os.path.join(export_folder, filename)
        
        wb.save(file_path)
        return file_path
    
    def export_game_to_excel(self, game_name: str) -> str:
        """导出指定游戏的数据到Excel"""
        try:
            # 获取游戏数据
            game_data = self.get_game_data(game_name)
            if not game_data:
                return f"❌ 未找到游戏: {game_name}"
            
            # 创建Excel文件
            file_path = self.create_excel_file(game_data)
            
            # 统计信息
            user_count = len(game_data['users'])
            total_records = sum(len(user['records']) for user in game_data['users'])
            completed_users = sum(1 for user in game_data['users'] if user['is_completed'])
            
            return f"✅ 导出成功!\n游戏: {game_name}\n用户数: {user_count}\n记录数: {total_records}\n完成用户: {completed_users}\n文件: {os.path.basename(file_path)}"
            
        except Exception as e:
            return f"❌ 导出失败: {str(e)}"
    
    def get_available_games(self) -> List[str]:
        """获取可用的游戏列表"""
        conn = sqlite3.connect(self.db_manager.db_path)
        cursor = conn.cursor()
        
        cursor.execute('SELECT name FROM games ORDER BY name')
        games = cursor.fetchall()
        conn.close()
        
        return [game[0] for game in games]
    
    def list_available_games(self) -> str:
        """列出可用的游戏"""
        games = self.get_available_games()
        
        if not games:
            return "❌ 数据库中没有游戏数据"
        
        game_list = []
        for game_name in games:
            # 获取统计信息
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT COUNT(DISTINCT u.id) as user_count, COUNT(r.id) as record_count
                FROM games g
                LEFT JOIN users u ON g.id = u.game_id
                LEFT JOIN records r ON u.id = r.user_id
                WHERE g.name = ?
            ''', (game_name,))
            
            result = cursor.fetchone()
            user_count = result[0] if result[0] else 0
            record_count = result[1] if result[1] else 0
            
            conn.close()
            
            game_list.append(f"• {game_name} ({user_count}用户, {record_count}记录)")
        
        return f"📁 可导出的游戏:\n" + "\n".join(game_list)
    
    def export_all_games(self) -> str:
        """导出所有游戏数据"""
        games = self.get_available_games()
        
        if not games:
            return "❌ 数据库中没有游戏数据"
        
        results = []
        success_count = 0
        
        for game_name in games:
            result = self.export_game_to_excel(game_name)
            if result.startswith("✅"):
                success_count += 1
            results.append(f"{game_name}: {'成功' if result.startswith('✅') else '失败'}")
        
        return f"📦 批量导出完成!\n成功: {success_count}/{len(games)}\n详情:\n" + "\n".join(results)
    
    def export_all_games_to_single_file(self) -> str:
        """将所有游戏导出到单个Excel文件的不同sheet中"""
        games = self.get_available_games()
        
        if not games:
            return "❌ 数据库中没有游戏数据"
        
        # 创建新的工作簿
        wb = Workbook()
        # 删除默认的sheet
        wb.remove(wb.active)
        
        success_count = 0
        failed_games = []
        
        for game_name in games:
            try:
                # 获取游戏数据
                game_data = self.get_game_data(game_name)
                if not game_data:
                    failed_games.append(f"{game_name}: 无数据")
                    continue
                
                # 创建新的工作表，使用游戏名作为sheet名
                # 确保sheet名符合Excel规范（最多31字符，不能包含特殊字符）
                safe_sheet_name = self._make_safe_sheet_name(game_name)
                ws = wb.create_sheet(title=safe_sheet_name)
                
                # 填充数据（复用create_excel_file的逻辑，但不保存文件）
                self._fill_worksheet_data(ws, game_data)
                
                success_count += 1
                
            except Exception as e:
                failed_games.append(f"{game_name}: {str(e)}")
        
        if success_count == 0:
            return f"❌ 所有游戏导出失败:\n" + "\n".join(failed_games)
        
        # 保存合并文件
        export_folder = os.path.join(self.config.excel_folder, "exports")
        os.makedirs(export_folder, exist_ok=True)
        
        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime("%m-%d-%H%M")
        filename = f"all_games_export_{timestamp}.xlsx"
        file_path = os.path.join(export_folder, filename)
        
        wb.save(file_path)
        
        result_lines = [f"📦 合并导出完成!"]
        result_lines.append(f"成功: {success_count}/{len(games)} 个游戏")
        result_lines.append(f"文件: {filename}")
        
        if failed_games:
            result_lines.append(f"失败的游戏:")
            result_lines.extend([f"  • {fail}" for fail in failed_games])
        
        return "\n".join(result_lines)
    
    def _make_safe_sheet_name(self, name: str) -> str:
        """将游戏名转换为安全的Excel sheet名"""
        # Excel sheet名限制：最多31字符，不能包含 : \ / ? * [ ]
        forbidden_chars = [':', '\\', '/', '?', '*', '[', ']']
        safe_name = name
        
        for char in forbidden_chars:
            safe_name = safe_name.replace(char, '_')
          # 限制长度为31字符
        if len(safe_name) > 31:
            safe_name = safe_name[:31]
        
        return safe_name
    
    def _fill_worksheet_data(self, ws, game_data: Dict):
        """填充工作表数据（从create_excel_file中提取的逻辑）"""
        game_name = game_data['game_name']
        users = game_data['users']
        
        if not users:
            # 如果没有用户数据，创建空表格
            # 不设置表头，直接设置行高
            ws.row_dimensions[1].height = self.config.row_height
        else:
            # 组织数据：将同名用户的不同周期合并显示
            organized_users = {}
            for user in users:
                base_name = user['name']
                cycle = user['cycle']
                
                if base_name not in organized_users:
                    organized_users[base_name] = {}
                
                organized_users[base_name][cycle] = {
                    'is_completed': user['is_completed'],
                    'records': user['records']                }
            
            # 不设置表头，从第1行开始显示用户数据
            current_row = 1
            
            # 处理每个用户
            for base_name, cycles in organized_users.items():
                for cycle, cycle_data in sorted(cycles.items()):                    # 用户名列
                    display_name = f"{base_name}({cycle})" if cycle > 1 else base_name
                    ws.cell(row=current_row, column=1, value=display_name)
                    ws.cell(row=current_row, column=1).alignment = self.center_alignment
                    ws.cell(row=current_row, column=1).fill = self.yellow_fill  # A列用户名设置为黄色
                    ws.row_dimensions[current_row].height = self.config.row_height
                    
                    is_completed = cycle_data['is_completed']
                    records = cycle_data['records']
                    
                    # 添加记录数据
                    col = 2
                    for record_date, count in records:
                        record_value = f"{record_date}_{count}"
                        ws.cell(row=current_row, column=col, value=record_value)
                        ws.cell(row=current_row, column=col).alignment = self.center_alignment
                        col += 1
                    
                    # 如果周期已完成，设置记录列（除A列外）为蓝色背景
                    if is_completed:
                        for c in range(2, col):  # 从第2列开始，跳过A列用户名
                            ws.cell(row=current_row, column=c).fill = self.blue_fill
                    
                    current_row += 1
        
        # 设置列宽
        ws.column_dimensions['A'].width = self.config.name_column_width
