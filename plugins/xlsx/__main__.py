from nonebot import on_command, get_driver
from nonebot.adapters.onebot.v11 import Message, MessageSegment, Bot, GroupMessageEvent, PrivateMessageEvent, MessageEvent
from nonebot.params import CommandArg
from nonebot.permission import SUPERUSER
from nonebot.exception import FinishedException
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment
import datetime
import os
import re
import glob
from pathlib import Path
from typing import Optional, Dict, Any
from .config import Config
from .database import DatabaseManager
from .excel_importer import ExcelImporter
from .excel_exporter import ExcelExporter

# 定义蓝色填充样式
BLUE_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

# 定义居中对齐样式
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

# 获取配置
plugin_config = Config()
# 初始化数据库管理器
db_manager = DatabaseManager()
# 初始化Excel导入器
excel_importer = ExcelImporter()
# 初始化Excel导出器
excel_exporter = ExcelExporter()
# 存储动态创建的命令处理器
command_handlers = {}

def find_latest_export_file(game_name: str) -> Optional[str]:
    """查找指定游戏的最新导出文件"""
    export_folder = os.path.join(plugin_config.excel_folder, "exports")
    
    if not os.path.exists(export_folder):
        return None
    
    # 查找匹配的文件模式: {game_name}_export_MM-DD-HHMM.xlsx
    pattern = f"{game_name}_export_*.xlsx"
    matching_files = glob.glob(os.path.join(export_folder, pattern))
    
    if not matching_files:
        return None
    
    # 按修改时间排序，获取最新的文件
    matching_files.sort(key=os.path.getmtime, reverse=True)
    return matching_files[0]

def get_games_from_database():
    """从数据库获取所有游戏名称"""
    games = db_manager.get_games_list()
    return [game[0] for game in games]

def register_game_commands():
    """基于数据库游戏表注册命令"""
    games = get_games_from_database()
    
    if plugin_config.debug_mode:
        print(f"数据库路径: {db_manager.db_path}")
        print(f"从数据库获取的游戏: {games}")
    
    if not games:
        print(f"⚠️  警告: 数据库中没有找到任何游戏")
        print(f"请先使用 /xlsximport 命令导入Excel文件到数据库")
        return
    
    for game_name in games:
        # 创建命令处理器（添加SUPERUSER权限）
        handler = on_command(game_name, priority=10, permission=SUPERUSER)
        
        # 创建处理函数的闭包，确保每个命令都有自己的game_name
        def create_handler(game_name):
            async def handler_func(args: Message = CommandArg()):
                result = await handle_excel_command(game_name, args)
                await handler.finish(result)
            return handler_func
        
        # 绑定处理函数
        handler.handle()(create_handler(game_name))
        
        # 存储处理器引用
        command_handlers[game_name] = handler
        
        if plugin_config.debug_mode:
            print(f"已注册命令: {game_name} -> 数据库存储")

async def handle_excel_command(game_name: str, args: Message = CommandArg()):
    """通用Excel命令处理函数 - 使用SQLite数据库"""
    cmd = args.extract_plain_text().strip()
    
    if not cmd:
        return f"❌ 命令格式错误！请使用以下格式：\n• /{game_name} <名字> +1\n• /{game_name} <名字> <次数>"
    
    # 解析命令格式
    # 支持格式：
    # 1. "用户名 +1" - 传统格式，添加1次记录
    # 2. "用户名 数字" - 新格式，添加指定次数的记录
    
    parts = cmd.split()
    if len(parts) < 2:
        return f"❌ 命令格式错误！请使用以下格式：\n• /{game_name} <名字> +1\n• /{game_name} <名字> <次数>"
    
    # 获取最后一部分作为次数参数
    count_part = parts[-1]
    username = " ".join(parts[:-1])  # 用户名可能包含空格
    
    # 解析次数
    if count_part == "+1":
        count = 1
    elif count_part.isdigit():
        count = int(count_part)
    else:
        return f"❌ 无效的次数格式！请使用 +1 或数字（如：/{game_name} {username} 5）"
    
    # 验证次数范围
    if count <= 0 or count > 100:
        return f"❌ 次数必须在1-100之间！"
    
    try:        # 添加用户记录
        result = db_manager.add_user_record(username, game_name, count)
        
        if count == 1:
            return f"✅ 已为 {username} 添加1次 {game_name} 记录\n{result}"
        else:
            return f"✅ 已为 {username} 添加{count}次 {game_name} 记录\n{result}"
            
    except Exception as e:
        return f"❌ 添加记录失败: {str(e)}"

# 注册文档导入命令
xlsximport_handler = on_command("文档导入", priority=5, permission=SUPERUSER)

@xlsximport_handler.handle()
async def handle_xlsximport(args: Message = CommandArg()):
    """处理Excel导入命令"""
    filename = args.extract_plain_text().strip()
    
    if not filename:
        # 如果没有指定文件名，列出可用文件
        result = excel_importer.list_available_files()
        await xlsximport_handler.finish(result)
    else:
        # 导入指定文件
        result = excel_importer.import_excel_file(filename)
        
        # 如果导入成功，重新注册游戏命令
        if result.startswith("✅"):
            register_game_commands()
        
        await xlsximport_handler.finish(result)

# 注册文档导出命令
xlsxexport_handler = on_command("文档导出", priority=5, permission=SUPERUSER)

@xlsxexport_handler.handle()
async def handle_xlsxexport(args: Message = CommandArg()):
    """处理Excel导出命令"""
    args_text = args.extract_plain_text().strip()
      # 检查是否包含 --upload 参数
    upload_file = "--upload" in args_text
    if upload_file:
        args_text = args_text.replace("--upload", "").strip()
    
    if not args_text:
        await xlsxexport_handler.finish("❌ 请提供游戏名称或使用 'all' 导出所有游戏\n使用方法: /文档导出 <游戏名|all> [--upload]")
    
    if args_text.lower() == "all":
        if upload_file:
            # 导出所有游戏并上传合并文件
            await handle_export_all_and_upload()
        else:
            # 使用合并导出功能，将所有游戏合并到一个Excel文件的不同sheet中
            result = excel_exporter.export_all_games_to_single_file()
            await xlsxexport_handler.finish(result)
    else:
        game_name = args_text
        if upload_file:
            # 导出指定游戏并上传文件
            await handle_export_and_upload(game_name)
        else:
            # 执行单个游戏导出
            result = excel_exporter.export_game_to_excel(game_name)
            await xlsxexport_handler.finish(result)

# 注册创建表格命令
xlsxcreate_handler = on_command("创建表格", priority=5, permission=SUPERUSER)

@xlsxcreate_handler.handle()
async def handle_xlsxcreate(args: Message = CommandArg()):
    """处理手动创建游戏命令"""
    game_name = args.extract_plain_text().strip()
    
    if not game_name:
        await xlsxcreate_handler.finish("❌ 请提供游戏名称！\n使用方法: /创建表格 <游戏名>")
    
    # 检查游戏名是否已存在
    existing_games = db_manager.get_games_list()
    existing_game_names = [game[0] for game in existing_games]
    
    if game_name in existing_game_names:
        await xlsxcreate_handler.finish(f"❌ 游戏 '{game_name}' 已存在！")
    
    try:
        # 添加新游戏到数据库
        game_id = db_manager.add_game(game_name)
        
        if game_id:
            # 重新注册命令以包含新创建的游戏
            register_game_commands()
            
            result_msg = f"✅ 成功创建游戏: {game_name}\n"
            result_msg += f"现在可以使用以下命令:\n"
            result_msg += f"• /{game_name} <用户名> +1 - 添加记录\n"
            result_msg += f"• /{game_name} <用户名> <次数> - 批量添加记录\n"
            result_msg += f"• /文档导出 {game_name} - 导出数据"
            
            await xlsxcreate_handler.finish(result_msg)
        else:
            await xlsxcreate_handler.finish(f"❌ 创建游戏失败: {game_name}")
            
    except Exception as e:
        await xlsxcreate_handler.finish(f"❌ 创建游戏时出错: {str(e)}")

# 注册表格查询命令
xlsxlookup_handler = on_command("表格查询", priority=5, permission=SUPERUSER)

@xlsxlookup_handler.handle()
async def handle_xlsxlookup(args: Message = CommandArg()):
    """处理查询用户记录命令"""
    args_text = args.extract_plain_text().strip()
    
    if not args_text:
        await xlsxlookup_handler.finish("❌ 请提供查询参数！\n使用方法: /表格查询 <游戏名> <用户名> [记录数量]")
    
    # 解析参数
    parts = args_text.split()
    if len(parts) < 2:
        await xlsxlookup_handler.finish("❌ 参数不足！\n使用方法: /表格查询 <游戏名> <用户名> [记录数量]")
    
    game_name = parts[0]
    username = parts[1]
    
    # 解析记录数量（可选参数）
    limit = plugin_config.default_lookup_count
    if len(parts) >= 3:
        try:
            limit = int(parts[2])
            if limit <= 0 or limit > 20:
                await xlsxlookup_handler.finish("❌ 记录数量必须在1-20之间！")
        except ValueError:
            await xlsxlookup_handler.finish("❌ 记录数量必须是数字！")
    
    try:
        # 获取用户摘要信息
        summary = db_manager.get_user_summary(username, game_name, limit)
        
        if "error" in summary:
            await xlsxlookup_handler.finish(f"❌ {summary['error']}")
        
        if not summary["has_records"]:
            await xlsxlookup_handler.finish(f"❌ 用户 '{username}' 在游戏 '{game_name}' 中没有记录")
        
        # 构建响应消息
        response_msg = f"📊 查询结果\n"
        response_msg += f"🎮 游戏: {summary['game_name']}\n"
        response_msg += f"👤 用户: {summary['username']}\n"
        response_msg += f"📈 当前进度: {summary['completion_progress']}\n"
        response_msg += f"📝 总记录数: {summary['total_count']}\n\n"
          # 显示最新记录
        response_msg += f"🕒 最新 {len(summary['latest_records'])} 条记录:\n"
        for i, (date, count) in enumerate(summary['latest_records'], 1):
            response_msg += f"{i}. {date} - 第{count}次\n"
        
        await xlsxlookup_handler.finish(response_msg)
        
    except FinishedException:
        # 重新抛出FinishedException，这是NoneBot的正常流程控制
        raise
    except Exception as e:
        await xlsxlookup_handler.finish(f"❌ 查询失败: {str(e)}")

# 注册xlsx帮助命令
xlsx_help_handler = on_command("xlsx帮助", priority=5, permission=SUPERUSER)

@xlsx_help_handler.handle()
async def handle_xlsx_help():
    """显示Excel插件帮助信息"""
    help_msg = "📚 Excel插件帮助\n\n"
    
    help_msg += "🎮 动态游戏指令:\n"
    help_msg += "• /<游戏名> <用户名> +1 - 添加1次记录\n"
    help_msg += "• /<游戏名> <用户名> <次数> - 添加指定次数记录\n"
    help_msg += "  例：/原神 张三 +1 或 /原神 张三 5\n\n"
    
    help_msg += "📁 文件管理指令:\n"
    help_msg += "• /文档导入 - 列出可导入的Excel文件\n"
    help_msg += "• /文档导入 <文件名> - 导入指定Excel文件\n"
    help_msg += "• /文档导出 <游戏名> - 导出指定游戏数据\n"
    help_msg += "• /文档导出 all - 导出所有游戏数据\n"
    help_msg += "• /文档导出 <游戏名|all> --upload - 导出并显示文件信息\n\n"
    
    help_msg += "🎯 游戏管理指令:\n"
    help_msg += "• /创建表格 <游戏名> - 创建新游戏并注册命令\n\n"
    
    help_msg += "📊 查询指令:\n"
    help_msg += "• /表格查询 <游戏名> <用户名> - 查询最新3条记录\n"
    help_msg += "• /表格查询 <游戏名> <用户名> <数量> - 查询指定数量记录\n\n"
    
    help_msg += "⚙️ 使用限制:\n"
    help_msg += "• 所有命令需要SUPERUSER权限\n"
    help_msg += "• 次数范围：1-100\n"
    help_msg += "• 查询记录数量范围：1-20\n"
    help_msg += "• 支持文件格式：.xlsx、.xls\n\n"
    
    help_msg += "💡 提示:\n"
    help_msg += "• 使用 /xlsx帮助 查看此帮助信息\n"
    help_msg += "• 游戏名会根据导入的Excel文件自动注册\n"
    help_msg += "• 达到完成次数后自动开始新周期"
    
    await xlsx_help_handler.finish(help_msg)

# ...existing code...

# 在插件加载时注册命令
driver = get_driver()

@driver.on_startup
async def startup():
    print("Excel插件正在启动...")
    print(f"配置的Excel目录: {plugin_config.excel_folder}")
    print(f"目录是否存在: {os.path.exists(plugin_config.excel_folder)}")
    
    # 如果目录不存在，尝试创建
    if not os.path.exists(plugin_config.excel_folder):
        try:
            os.makedirs(plugin_config.excel_folder, exist_ok=True)
            print(f"已创建目录: {plugin_config.excel_folder}")
        except Exception as e:
            print(f"创建目录失败: {e}")
    
    # 基于数据库注册游戏命令
    register_game_commands()
    
    if len(command_handlers) == 0:
        print("⚠️  没有注册任何命令!")
        print("解决方案:")
        print("1. 使用 /xlsximport 命令导入Excel文件到数据库")
        print("2. 或者手动在数据库中添加游戏数据")
        print("3. 命令将在有游戏数据后自动可用")
    else:
        print(f"✅ Excel插件启动完成，已注册 {len(command_handlers)} 个命令")
        if plugin_config.debug_mode:
            print("注册的命令列表:", list(command_handlers.keys()))

@driver.on_shutdown
async def shutdown():
    print("Excel插件已关闭")

async def upload_file_to_chat(file_path: str, filename: Optional[str] = None) -> Message:
    """显示文件导出信息"""
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        # 如果没有指定文件名，使用原文件名
        if filename is None:
            filename = os.path.basename(file_path)
        
        # 获取文件信息
        file_size = os.path.getsize(file_path)
        file_size_mb = file_size / (1024 * 1024)
        
        # 构建文件信息消息
        message = Message()
        message += MessageSegment.text(f"📎 文件导出完成\n")
        message += MessageSegment.text(f"📁 文件名: {filename}\n")
        message += MessageSegment.text(f"📊 大小: {file_size_mb:.2f}MB ({file_size:,} bytes)\n")
        message += MessageSegment.text(f"💾 保存路径: {file_path}\n")
        message += MessageSegment.text(f"💡 请从服务器获取Excel文件")
        
        return message
        
    except Exception as e:
        raise Exception(f"文件处理失败: {str(e)}")

async def handle_export_and_upload(game_name: str):
    """导出指定游戏并上传文件"""
    try:
        # 先导出文件
        result = excel_exporter.export_game_to_excel(game_name)
        
        if not result.startswith("✅"):
            await xlsxexport_handler.finish(result)
            return
        
        # 查找最新的导出文件
        file_path = find_latest_export_file(game_name)
        
        if not file_path:
            await xlsxexport_handler.finish(f"❌ 未找到 {game_name} 的导出文件")
            return
        
        filename = os.path.basename(file_path)
        
        # 上传文件
        file_message = await upload_file_to_chat(file_path, filename)
        
        # 发送结果消息和文件
        await xlsxexport_handler.send(f"📤 {result}")
        await xlsxexport_handler.finish(file_message)
        
    except FinishedException:
        # 重新抛出FinishedException，这是NoneBot的正常流程控制
        raise
    except Exception as e:
        await xlsxexport_handler.finish(f"❌ 导出上传失败: {str(e)}")

async def handle_export_all_and_upload():
    """导出所有游戏并上传合并文件"""
    try:
        # 使用合并导出功能，将所有游戏合并到一个Excel文件的不同sheet中
        result = excel_exporter.export_all_games_to_single_file()
        
        if not result.startswith("📦"):
            await xlsxexport_handler.finish(result)
            return
        
        # 查找最新的合并导出文件
        export_folder = os.path.join(plugin_config.excel_folder, "exports")
        
        if not os.path.exists(export_folder):
            await xlsxexport_handler.finish("❌ 导出目录不存在")
            return
        
        # 查找匹配的合并文件模式: all_games_export_MM-DD-HHMM.xlsx
        import glob
        pattern = "all_games_export_*.xlsx"
        matching_files = glob.glob(os.path.join(export_folder, pattern))
        
        if not matching_files:
            await xlsxexport_handler.finish("❌ 未找到合并导出文件")
            return
        
        # 按修改时间排序，获取最新的文件
        matching_files.sort(key=os.path.getmtime, reverse=True)
        file_path = matching_files[0]
        filename = os.path.basename(file_path)
        
        # 发送结果消息
        await xlsxexport_handler.send(f"📤 {result}")
        
        # 上传合并文件
        file_message = await upload_file_to_chat(file_path, filename)
        await xlsxexport_handler.finish(file_message)
        
    except FinishedException:
        # 重新抛出FinishedException，这是NoneBot的正常流程控制
        raise
    except Exception as e:
        await xlsxexport_handler.finish(f"❌ 合并导出上传失败: {str(e)}")
